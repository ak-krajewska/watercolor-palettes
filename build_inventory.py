#!/usr/bin/env python3
"""
Rebuild the paints table in data/paints.db from source files.

Run this whenever you download a fresh paints-inventory.xlsx from artistpigments.org.
Manual additions (paints-manual-notes.csv) are merged in automatically.

On first run (or with --seed-from-csv), the containers, palettes, and loadouts
tables are seeded from the CSV files in data/. After that, those tables are
hand-curated in the database and not overwritten.

Usage:
    python3 build_inventory.py              # rebuild paints table only
    python3 build_inventory.py --seed-from-csv  # also (re)seed curated tables from CSVs
"""

import openpyxl, csv, re, sqlite3, sys
from pathlib import Path

BASE = Path(__file__).parent
XLSX = BASE / 'paints-inventory.xlsx'
MANUAL_CSV = BASE / 'paints-manual-notes.csv'
PIGMENT_INDEX = BASE / 'pigment-index.csv'
DB = BASE / 'data' / 'paints.db'
PALETTES_CSV = BASE / 'data' / 'palettes.csv'
CONTAINERS_CSV = BASE / 'data' / 'containers.csv'
LOADOUTS_CSV = BASE / 'data' / 'loadouts.csv'

PAINTS_FIELDS = ['id', 'color_name', 'brand', 'manufacturer_code', 'medium', 'hue_category',
                 'transparency', 'granulation', 'staining', 'lightfastness', 'astm_lightfastness',
                 'pigments', 'single_pigment', 'pigment_known', 'alt_names', 'source', 'notes']


def norm_name(s):
    """Normalize color name: strip leading 'CfM ' brand prefix, lowercase, alphanumeric only."""
    s = re.sub(r'^cfm\s+', '', s.strip(), flags=re.IGNORECASE)
    return re.sub(r'[^a-z0-9]', '', s.lower()) if s else ''

BRAND_MAP = {
    'cfmhandmadewatercolors': 'cfm',
    'winsornewtonprofessionalwatercolour20132': 'winsornewton',
    'winsornewtonprofessionalwatercolour2013': 'winsornewton',
    'winsornewtonprofessionalwatercolour': 'winsornewton',
    'winsornewtonprofessionalwatercolours': 'winsornewton',
    'davinciwatercolors': 'davinci',
    'winsornewtoncotwanwatercolours': 'winsornewton',
    'winsornewtoncotwanwatercolors': 'winsornewton',
    'winsornewton': 'winsornewton',
    'holbeinartistswatercolorhwc': 'holbein',
    'holbeinartistswatercolourhwc': 'holbein',
    'holbeinartistsgouache': 'holbein',
    'romanszmalaquarius': 'romanszmal',
    'cfm': 'cfm',
    'davinci': 'davinci',
}

def norm_brand(s):
    n = re.sub(r'[^a-z0-9]', '', s.lower()) if s else ''
    # Try progressively shorter prefixes to catch versioned brand strings
    for key in BRAND_MAP:
        if n.startswith(key) or n == key:
            return BRAND_MAP[key]
    return n


def load_xlsx():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb['Collection']
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    paints = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        pigments = [d[f'Pigment Code #{i}'] for i in range(1, 11) if d.get(f'Pigment Code #{i}')]
        paints.append({
            'id': d['ID'],
            'color_name': d['Color Name'],
            'brand': d['Brand'],
            'manufacturer_code': re.sub(r'[^\w]', '', str(d['Code'])) if d.get('Code') else '',
            'medium': d['Medium'],
            'transparency': d['Transparency'] or '',
            'granulation': d['Granulation'] or '',
            'staining': d['Staining'] or '',
            'lightfastness': d['Lightfastness'] or '',
            'astm_lightfastness': d['ASTM Lightfastness'] or '',
            'pigments': ', '.join(pigments),
            'single_pigment': 'yes' if len(pigments) == 1 else ('no' if len(pigments) > 1 else ''),
            'hue_category': '',
            'pigment_known': 'yes' if pigments else 'no',
            'alt_names': '',
            'source': 'artistpigments',
            'notes': d['Comment'] or '',
        })
    return paints


def load_manual():
    with open(MANUAL_CSV) as f:
        return list(csv.DictReader(f))


def load_pigment_index():
    """Build a lookup from individual pigment code -> color family.
    Only uses single-pigment rows (no '+') as the canonical family for a pigment.
    """
    lookup = {}
    with open(PIGMENT_INDEX) as f:
        for row in csv.DictReader(f):
            code = row['Color Index Name'].strip()
            family = row['family'].strip()
            if code and family and '+' not in code:
                lookup[code.upper()] = family
    return lookup


def assign_hue_categories(paints, pigment_index):
    """Set hue_category from the pigment index using the first known pigment code."""
    for p in paints:
        if p['hue_category']:
            continue  # already set from manual notes
        if not p['pigments']:
            continue
        first_pigment = p['pigments'].split(',')[0].strip().upper()
        if first_pigment in pigment_index:
            p['hue_category'] = pigment_index[first_pigment]


def merge(paints, manual_rows):
    # Build lookup by (norm_name, norm_brand)
    lookup = {}
    for m in manual_rows:
        key = (norm_name(m['Color name']), norm_brand(m['Brand']))
        lookup[key] = m

    for p in paints:
        key = (norm_name(p['color_name']), norm_brand(p['brand']))
        if key in lookup:
            m = lookup[key]
            p['hue_category'] = m.get('Hue', '')
            p['alt_names'] = m.get('Alternative names/similar names', '')
            # Fill in pigment data from manual if missing
            if not p['pigments'] and m.get('Pigment 1'):
                pigs = [m[f'Pigment {i}'] for i in range(1, 4) if m.get(f'Pigment {i}')]
                p['pigments'] = ', '.join(pigs)
                p['single_pigment'] = 'yes' if len(pigs) == 1 else 'no'
                p['pigment_known'] = 'yes'

    # Add paints that are in manual but not in xlsx (uncatalogued on artistpigments)
    catalogued = {norm_name(p['color_name']) for p in paints}
    extra_id = 1
    added = []
    for m in manual_rows:
        if norm_name(m['Color name']) not in catalogued and norm_brand(m['Brand']) == 'cfm':
            pigs = [m[f'Pigment {i}'] for i in range(1, 4) if m.get(f'Pigment {i}')]
            paints.append({
                'id': f'cfm_x{extra_id:02d}',
                'color_name': m['Color name'],
                'brand': 'CfM Handmade Watercolors',
                'manufacturer_code': '',
                'medium': 'Watercolor',
                'transparency': '',
                'granulation': '',
                'staining': '',
                'lightfastness': '',
                'astm_lightfastness': '',
                'pigments': ', '.join(pigs),
                'single_pigment': 'yes' if len(pigs) == 1 else ('no' if len(pigs) > 1 else ''),
                'hue_category': m.get('Hue', ''),
                'pigment_known': 'yes' if pigs else 'no',
                'alt_names': m.get('Alternative names/similar names', ''),
                'source': 'manual',
                'notes': '',
            })
            added.append(m['Color name'])
            extra_id += 1

    return paints, added


def init_db(conn):
    """Create all tables if they don't exist."""
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS paints (
            id TEXT PRIMARY KEY,
            color_name TEXT NOT NULL,
            brand TEXT,
            manufacturer_code TEXT,
            medium TEXT,
            hue_category TEXT,
            transparency TEXT,
            granulation TEXT,
            staining TEXT,
            lightfastness TEXT,
            astm_lightfastness TEXT,
            pigments TEXT,
            single_pigment TEXT,
            pigment_known TEXT,
            alt_names TEXT,
            source TEXT,
            notes TEXT
        );

        CREATE TABLE IF NOT EXISTS containers (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            brand TEXT,
            max_slots INTEGER,
            portability TEXT,
            pan_orientation TEXT,
            notes TEXT
        );

        CREATE TABLE IF NOT EXISTS palette_names (
            name TEXT PRIMARY KEY
        );

        CREATE TABLE IF NOT EXISTS palettes (
            palette_name TEXT NOT NULL REFERENCES palette_names(name),
            paint_id TEXT NOT NULL REFERENCES paints(id),
            color_name TEXT,
            row TEXT,
            position INTEGER,
            notes TEXT,
            PRIMARY KEY (palette_name, paint_id)
        );

        CREATE TABLE IF NOT EXISTS loadouts (
            palette_name TEXT NOT NULL REFERENCES palette_names(name),
            container_id TEXT NOT NULL REFERENCES containers(id),
            PRIMARY KEY (palette_name)
        );
    ''')


def write_paints(conn, paints):
    """Delete and rebuild the paints table from source data."""
    conn.execute('DELETE FROM paints')
    placeholders = ', '.join(['?'] * len(PAINTS_FIELDS))
    conn.executemany(
        f'INSERT INTO paints ({", ".join(PAINTS_FIELDS)}) VALUES ({placeholders})',
        [[p[f] for f in PAINTS_FIELDS] for p in paints]
    )


def seed_from_csv(conn):
    """Seed the curated tables (containers, palettes, loadouts) from CSV files."""
    seeded = []

    if CONTAINERS_CSV.exists():
        conn.execute('DELETE FROM containers')
        with open(CONTAINERS_CSV) as f:
            rows = list(csv.DictReader(f))
        for row in rows:
            conn.execute(
                'INSERT INTO containers (id, name, brand, max_slots, portability, pan_orientation, notes) '
                'VALUES (?, ?, ?, ?, ?, ?, ?)',
                (row['id'], row['name'], row['brand'],
                 int(row['max_slots']) if row.get('max_slots') else None,
                 row.get('portability', ''), row.get('pan_orientation', ''), row.get('notes', ''))
            )
        seeded.append(f'containers ({len(rows)})')

    if PALETTES_CSV.exists():
        conn.execute('DELETE FROM palettes')
        conn.execute('DELETE FROM palette_names')
        with open(PALETTES_CSV) as f:
            rows = list(csv.DictReader(f))
        # Extract and insert distinct palette names first
        names = sorted(set(row['palette_name'] for row in rows))
        for name in names:
            conn.execute('INSERT INTO palette_names (name) VALUES (?)', (name,))
        for row in rows:
            conn.execute(
                'INSERT INTO palettes (palette_name, paint_id, color_name, row, position, notes) '
                'VALUES (?, ?, ?, ?, ?, ?)',
                (row['palette_name'], row['paint_id'], row.get('color_name', ''),
                 row.get('row', ''),
                 int(row['position']) if row.get('position') else None,
                 row.get('notes', ''))
            )
        seeded.append(f'palette_names ({len(names)}), palettes ({len(rows)})')

    if LOADOUTS_CSV.exists():
        conn.execute('DELETE FROM loadouts')
        with open(LOADOUTS_CSV) as f:
            rows = list(csv.DictReader(f))
        for row in rows:
            conn.execute(
                'INSERT INTO loadouts (palette_name, container_id) VALUES (?, ?)',
                (row['palette_name'], row['container_id'])
            )
        seeded.append(f'loadouts ({len(rows)})')

    return seeded


def tables_empty(conn):
    """Check if the curated tables have no data (need initial seeding)."""
    for table in ('containers', 'palette_names', 'palettes', 'loadouts'):
        count = conn.execute(f'SELECT COUNT(*) FROM {table}').fetchone()[0]
        if count > 0:
            return False
    return True


if __name__ == '__main__':
    do_seed = '--seed-from-csv' in sys.argv

    paints = load_xlsx()
    manual = load_manual()
    pigment_index = load_pigment_index()
    paints, added = merge(paints, manual)
    assign_hue_categories(paints, pigment_index)
    for p in paints:
        if p['hue_category']:
            p['hue_category'] = p['hue_category'].title()

    conn = sqlite3.connect(DB)
    conn.execute('PRAGMA foreign_keys = ON')
    init_db(conn)

    # Write paints first -- palettes reference them via foreign key
    write_paints(conn, paints)

    # Seed curated tables on first run or when explicitly requested
    if do_seed or tables_empty(conn):
        seeded = seed_from_csv(conn)
        if seeded:
            print(f"Seeded from CSV: {', '.join(seeded)}")

    conn.commit()

    count = conn.execute('SELECT COUNT(*) FROM paints').fetchone()[0]
    print(f"Wrote {count} paints to {DB}")
    if added:
        print(f"Added from manual (not on artistpigments): {', '.join(added)}")
    missing_hue = conn.execute(
        "SELECT COUNT(*) FROM paints WHERE hue_category = '' OR hue_category IS NULL"
    ).fetchone()[0]
    print(f"Paints without hue category: {missing_hue}")

    # Check foreign key integrity
    fk_errors = conn.execute('PRAGMA foreign_key_check').fetchall()
    if fk_errors:
        print(f"WARNING: {len(fk_errors)} foreign key violations found:")
        for table, rowid, ref_table, fk_idx in fk_errors:
            print(f"  {table} rowid={rowid} -> {ref_table}")
    else:
        print("Foreign key references OK.")

    conn.close()
